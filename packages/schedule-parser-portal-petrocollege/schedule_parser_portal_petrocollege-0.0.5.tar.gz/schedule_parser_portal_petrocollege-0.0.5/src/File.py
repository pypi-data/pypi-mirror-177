import openpyxl
from datetime import datetime

from Lesson import Lesson


class ExcelFile:

    def __init__(self, path_file):
        self.file = openpyxl.load_workbook(path_file)
        self.active_sheet = self.file.active
        self.dates = self.active_sheet['A']

    def get_dates_with_rows_number(self):
        """
        Return list of dates lessons with rows excel number
        :return: list [{'date': datetime.datetime(2022, 9, 1, 0, 0), 'rows': [20, 21, 22, 23, 24]}]
        """
        list_dates_rows_number = []
        for x in self.dates:
            if x.value is not None and any(map(str.isdigit, x.value)):
                list_dates_rows_number.append({
                    "date": self.__split_date(x.value),
                    "rows": list(range(x.row, x.row + 5))
                })
        return list_dates_rows_number

    def __split_date(self, text_date):
        """
        Get data from dates schedule
        :param text_date: string  ex. Чт 01.09.22
        :return: datetime.datetime(2022, 9, 1, 0, 0)
        """
        date = datetime.strptime(text_date.split(" ")[-1], "%d.%m.%y")
        return date

    def get_object(self):
        """
        :return: dict
        {
            'teacher': 'Ярошенко С.П.',
            'debug_column': 324,
            'teacher_lessons':
                [
                    {
                        'lesson':
                            {
                                'discipline': 'Теор.гос.и права',
                                'room': '101',
                                'building': '4',
                                'groups': ['11-29'],
                                 '_Lesson__current_string': 'Теор.гос.и права  4/101',
                                'is_dop': False,
                                'subgroup': 0
                            },
                            'date_lesson': datetime.datetime(2022, 9, 1, 0, 0),
                            'number_of_lesson': 2,
                            'debug_position_row': 21,
                            'debug_position_column': 324,
                            'debug_position_coordinate': 'LL21'
                            }
                    }
                    ...
                ]
        }
        """
        teachers_has_lessons = []
        date_lesson = datetime.min
        dates_with_rows_number = self.get_dates_with_rows_number()
        cols = self.active_sheet.max_column - 1
        for col in self.active_sheet.iter_cols(min_col=3, min_row=1, max_col=cols):
            teacher_has_lesson = {}
            for cell in col:
                if cell.row == 1:
                    teacher_has_lesson['teacher'] = cell.value
                    teacher_has_lesson['debug_column'] = cell.column
                    teacher_has_lesson["teacher_lessons"] = []
                elif cell.value is not None and cell.value.strip():
                    for item in dates_with_rows_number:
                        if cell.row in item['rows']:
                            date_lesson = item['date']  # Дата проведения занятия
                    teacher_has_lesson["teacher_lessons"].append(
                        {
                            "lesson": Lesson(cell.value).__dict__,
                            'date_lesson': date_lesson,
                            "number_of_lesson": self.active_sheet["B" + str(cell.row)].value,
                            "debug_position_row": cell.row,
                            "debug_position_column": cell.column,
                            "debug_position_coordinate": cell.coordinate,
                        })
            teachers_has_lessons.append(teacher_has_lesson)
        return teachers_has_lessons
