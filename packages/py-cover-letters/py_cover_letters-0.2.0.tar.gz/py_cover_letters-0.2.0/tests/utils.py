from pathlib import Path
from typing import List, Dict, Any

from openpyxl.reader.excel import load_workbook


def read_excel(filename: Path, sheet_name: str, column_mapping: Dict[int, str],
               include_headers: bool = False) -> List[Dict[str, Any]]:
    cover_letters = list()
    wb = load_workbook(filename)
    sheet = wb[sheet_name]
    last_row = sheet.max_row + 1

    start_at = 1 if include_headers else 2
    for row in range(start_at, last_row):
        cover_letter_dict = dict()
        for col, name in column_mapping.items():
            cell_obj = sheet.cell(row=row, column=col)
            value = cell_obj.value
            cover_letter_dict[name] = value
        cover_letters.append(cover_letter_dict)
    return cover_letters
