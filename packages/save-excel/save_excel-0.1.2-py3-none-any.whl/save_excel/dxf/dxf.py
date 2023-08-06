from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font

class Dxf:
    """
    Contains attributes with template styles for conditional format.
    
    Examples
    ----------
    Calling this class:
    from openpyxl.format.rule import Rule, IconSetRule, ColorScaleRule, FormulaRule, CellIsRule, DataBarRule
    dxf = Dxf()
    
    Creating some rules using this template styles:
    red_rule = Rule(type="cellIs", operator="lessThan", dxf=dxf.red, formula=['25'])
    yellow_rule = Rule(type="cellIs", operator="between", dxf=dxf.yellow, formula=['25', '50'])
    green_rule = Rule(type="cellIs", operator="greaterThan", dxf=dxf.green, formula=['50'])
    blue_rule = Rule(type="containsText", operator="containsText", dxf=dxf.blue, formula=['NOT(ISERROR(SEARCH("TEXTO",A1)))'])
    orange_rule = Rule(type="containsText", operator="containsText", dxf=dxf.orange, formula=['NOT(ISERROR(SEARCH("TEXTO",{})))'])
    top_rule = Rule(type="top10", dxf=dxf.red, rank=5)
    
    Another rules:
    icon_rule = IconSetRule('3TrafficLights2', type='percent', values=[0, 50, 100], reverse=False, showValue=False)
    color_scale_rule = ColorScaleRule(start_type='min', start_value=None, start_color='FF8181', end_type='max', end_value=None, end_color='7BFF71')
    color_scale_rule = ColorScaleRule(start_type='num', start_value=0, start_color='FF8181', mid_type='num', mid_value=50, mid_color='FFFF57', end_type='num', end_value=100, end_color='7BFF71')
    databar_rule = DataBarRule(start_type='min', start_value=None, end_type='max', end_value=None, color='FFC7CE', showValue=True, minLength=10, maxLength=None)
    """
    
    def __init__(self):
        self.red = DifferentialStyle(font=Font(color="9C0006"), fill=PatternFill(bgColor="FFC7CE"))
        self.yellow = DifferentialStyle(font=Font(color="9C6500"), fill=PatternFill(bgColor="FFEB9C"))
        self.green = DifferentialStyle(font=Font(color="006100"), fill=PatternFill(bgColor="C6EFCE"))
        self.cyan = DifferentialStyle(font=Font(color="FFFFFF"), fill=PatternFill(bgColor="008B8B"))
        self.orange = DifferentialStyle(font=Font(color="000000"), fill=PatternFill(bgColor="F58D16"))