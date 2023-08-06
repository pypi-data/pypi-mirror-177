"""
ExcelExtractor is an module to handle Excel documents in a normal table format for python
"""

import copy
import json
import re
import warnings

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import cell
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings('ignore')


class ExcelExtractor:
    """
    ExcelExtractor handles maschine readable data for excel documents

    ExcelExtrator handles excel worksheet with an headline and data under it.
    It can read/write from those data and presents them in an maschine readable format in python
    Needed file Structure in Excel:
         |    A    |    B    |    C    | ....
         +---------+---------+---------+-------
         | ... 
         +---------+---------+---------+-------
         | Header1 | Header2 | Header3 | ...
         +---------+---------+---------+-------
         | Content | Content | Content | ...
         +---------+---------+---------+-------
         | ...
    """

    def __init__(self, document, newFile=False):
        """
        Constructs the ExcelExtractor class
        
        Constructor loads the excel document. Only xlsx is supported
        
        :param document: input file with path
        :type document: string
        :param newFile: If true a new file will be generated
        :type newFile: bool
        :raises PermissionError: Raises if the file can't be open due to missing permissions
        :raises FileNotFoundError: Raises if the file can't be found
        :raises ValueError: Raises if the file is not in the correct format
        """
        self.file = document
        if newFile is True:
            self.doc = Workbook()
        else:
            try:
                self.doc = load_workbook(filename = document)
            except PermissionError:
                raise PermissionError('Error: Can\'t open file. Permission denied: \'%s\'' % self.file)
            except FileNotFoundError:
                raise FileNotFoundError('Error: Can\'t open file. Not found: \'%s\'' % self.file)
            except InvalidFileException:
                raise ValueError('Given file format is not supported. Supported formats are: .xlsx,.xlsm,.xltx,.xltm')
        self.sheetTitle = None
        self.sheet = None
        self.header = { }

        self.validators = { }

        self.headerRow = None
        self.lastColumn = 0

        self.warnings = []

    def setSheetFromName(self, name):
        """
        Sets the worksheet from the sheetname
        
        Uses the input name to set the worksheet for the data extraction. Raises an error if it fails
        
        :param name: Excel worksheet name
        :type name: string
        :raises KeyError: Raises if the worksheetname is not found
        """
        if name in self.doc.sheetnames:
            self.sheet = self.doc[name]
            self.sheetTitle = name
        else:
            raise KeyError('Sheetname \'%s\' not found' % name)

    def setSheetFromId(self, sheetID):
        """
        Sets the worksheet from the sheet ID
        
        Uses the input ID to set the worksheet for the data extraction. Raises an error if it fails
        
        :param name: Excel worksheet ID
        :type name: int
        :raises KeyError: Raises if the worksheet ID is not found
        """
        if sheetID < len(self.doc.sheetnames):
            self.sheetTitle = self.doc.sheetnames[sheetID]
            self.sheet = self.doc[self.sheetTitle]
        else:
            raise KeyError('SheetID \'%s\' not found' % sheetID)

    def findHeaderRow(self):
        """
        Finds the header row of the worksheet
        
        Finds the header row of the worksheet. It searchs in column A for a cell with an defined header in it.
        The headers are stored in self.header
        
        :raises ValueError: Raises if no header can be found
        """
        if self.headerRow is None:
            for cell in self.sheet['A']:
                if cell.value is not None and str(cell.value).strip().lower() in self.header.keys():
                    self.headerRow = cell.row
                    return
            raise ValueError('Header row not found')

    def addHeader(self, header, clean=False, default='', errorLevel=0, width=0):
        """
        Adds an header to the header list
        
        The set header are the headers where ExcelExtractor'll return values from.
        Here you can define cleaning patterns, default values and an errorLevel if the value in a cell is missing.
        Error level:
            0: No Error
            1: Raises an Warning
            2: Raises an ValueError
        
        :param header: Header name
        :type header: string
        :param clean: When set to true the cell value will be cleaned with the pattern defined in addCleanPatternToHeader(), defaults to False
        :param clean: bool, optional
        :param default: default value of a empty cell, defaults to ''
        :param default: str, optional
        :param errorLevel: Error level for this cell value, defaults to 0
        :param errorLevel: int, optional
        :param width: Width of the column of the specific header, defaults to 0 (no width change)
        :param width: int, optional
        :raises ValueError: Raises if the header is already set
        """
        if header.strip().lower() not in self.header.keys():
            self.header[header.strip().lower()] = [header, None, clean, default, errorLevel, [], width] # @TODO: Check params
        else:
            raise ValueError('%s already as header set' % header)

    def addCleanPatternToHeader(self, header, pattern, replace):
        """
        Adds an regex pattern to clean cell values
        
        The added regex pattern cleans the cell value. You can define here more than one pattern for each header.
        
        :param header: Name of the header
        :type header: string
        :param pattern: Regex pattern for the replacement
        :type pattern: string
        :param replace: Replacement for the pattern
        :type replace: string
        :raises ValueError: Raises if the pattern is not a valid regex pattern  
        :raises KeyError: Raises if the header is not in the header list
        """
        if header.strip().lower() in self.header.keys():
            try:
                re.compile(pattern)
                self.header[header.strip().lower()][5].append((pattern, replace))
            except re.error:
                raise ValueError('\'%s\' is not a valid regex pattern' % pattern)
        else:
            raise KeyError('%s not in header list' % header) 
                
    def findHeaderColumns(self, append = False):
        """
        Finds the header columns for all headers in the header list
        
        First the header row is searched, after that the function iterates through is row to find the header columns
        
        :param append: If set to True the missing headers will be appended, defaults to False
        :param append: bool, optional
        :raises ValueError: Raises if there are more than one header with the same name in the document
        :raises ValueError: Raises if not all header are found and append is false
        """
        self.findHeaderRow()
        for x in self.sheet[self.headerRow]:
            if x.value is not None and x.value.strip().lower() in self.header.keys():
                if self.header[x.value.strip().lower()][1] is None:
                    self.header[x.value.strip().lower()][1] = x
                else:
                    raise ValueError('Header %s more than once in the %s' % (x.value, self.file))
        missingHeader = [] # @TODO: Fix last column
        
        for k,v in self.header.items():
            if v[1] is None:
                missingHeader.append(v[0])

        if(append == True):
            self.appendHeaders(missingHeader)
        elif len(missingHeader) != 0:
            raise ValueError('Not all headers found in the document: %s' % missingHeader)

    """
    Appends missing headers to the header row
    
    :param missingHeaders: List of missing headers
    :param missingHeaders: list, required
    """    
    def appendHeaders(self, missingHeaders):
        self.findHeaderRow()
        self.lastColumn = self.sheet.max_column
        if(len(missingHeaders) == len(self.header)):
            self.lastColumn = 0

        for v in missingHeaders:
            self.lastColumn += 1
            self.header[v.lower()][1] = self.sheet.cell(row=self.headerRow, column=self.lastColumn, value=v)
            self.header[v.lower()][1].font = Font(size=13, bold=True)
            self.header[v.lower()][1].alignment = Alignment(wrapText=True)
        

    def save(self):
        """
        Saves the edited file
        
        Saves the file to the disk. This overwrites the original file.
        Also adds the filter in the headline and the validators.
        
        :raises PermissionError: Raises if the file can't be written
        """
        self.addFilter()
        self.setColumnWidth()
        for headerID in self.validators:
            self.addValidatorToCells(headerID)
        try:
            self.doc.save(self.file)
        except PermissionError:
            raise PermissionError('Error: Can\'t save file. Permission denied: \'%s\'' % self.file)
        
    def getCellValue(self, cell, clean=False, default='', errorLevel=0, patterns=[]):
        """
        Gets the value of a cell
        
        Gets the cell value from the given cell. Cleans up the cell if configured and sets an default
        value if the cell is empty. If an error level is set an warning or error will be raised.
        Error level:
            0: No Error
            1: Raises an Warning
            2: Raises an ValueError
        
        :param cell: Cell object from the worksheet
        :type cell: openpyxl.cell.cell
        :param clean: If True cleans the cell with patterns, defaults to False
        :param clean: bool, optional
        :param default: Default value for the cell, defaults to ''
        :param default: str, optional
        :param errorLevel: Error level if the cell is empty, defaults to 0
        :param errorLevel: int, optional
        :param patterns: Regex pattern for the cleanup, defaults to []
        :param patterns: list, optional
        :raises ValueError: Raises if the error level is 2 and the cell is empty
        :return: cell value
        :rtype: string
        """
        if cell.value is None:
            if errorLevel == 2:
                raise ValueError('Unhandled empty cell at: %s' % cell.coordinate)
            elif errorLevel == 1:
                self.warnings.append('Empty cell at %s filled with default: \'%s\'' % (cell.coordinate, default))
            return str(default)
        else:
            if clean is True:
                return self.cleanCell(str(cell.value).strip(), patterns)
            else:
                return str(cell.value).strip()

    def cleanCell(self, cell, patterns):
        """
        Cleans the cell with an regex pattern
        
        Cleans the cell with an dict of patterns
        
        :param cell: cell value
        :type cell: string
        :param patterns: dict of patterns with key: value of pattern: replacement
        :type patterns: dict
        :return: cleaned cell
        :rtype: string
        """
        for pattern in patterns:
            cell =  re.sub(pattern[0],pattern[1],str(cell))
        return cell

    def getData(self):
        """
        Gets the data of the worksheet.
        
        Gets the data of the worksheet as an list of dicts.
        A dict contains a key: value pair with headerName: cellValue
        
        :return: Data of the worksheet
        :rtype: list
        """
        rows = []
        for row in self.sheet:
            if row[0].row > self.headerRow:
                rows.append(self.getRow(row))
        rows = [x for x in rows if x is not None]
        return rows

    def getRow(self, row):
        """
        Get all data of a single row
        
        Returns all data of the given row.
        
        :param row: Row to excract the data from
        :type row: tupel
        :return: Dict of the processed data
        :rtype: dict
        """
        o = { }
        filled = False
        for v in self.header.values():
            if row[v[1].col_idx-1].value is not None and row[v[1].col_idx-1].value != '':
                filled = True
                break
        if filled is True:
            for k,v in self.header.items():
                o[v[0]] = self.getCellValue(row[v[1].col_idx-1], v[2], v[3], v[4], v[5])
        else:
            o = None
        return o

    def getRowData(self, rowID):
        """
        Get all data of a single row with the given ID
        
        Returns all data of the given row ID.
        
        :param row: Row ID to excract the data from
        :type row: tupel
        :return: Dict of the processed data
        :rtype: dict
        """
        row = self.sheet[rowID+self.headerRow+1]
        o = self.getRow(row)
        if o is None:
            o = {}
            for k,v in self.header.items():
                o[v[0]] = None
        return o

    def getRowFromID(self, rowID):
        """
        Gets the source row from an row ID
        
        Gets the source row from the given row ID.
        
        :param rowID: Row id. Relative to the header row. Starts at 0.
        :type rowID: int
        :raises KeyError: Raises if the row ID is not valid
        :return: Raw source row
        :rtype: tupel
        """
        if rowID >= 0:
            return self.sheet[rowID+self.headerRow+1]
        else:
            raise KeyError('Boundary exception for row ID: %s' % rowID)

    def getRowFromRealRowID(self, rowID):
        """
        Gets raw source row data with an absolute ID
                
        :param rowID: Absolute row ID
        :type rowID: int
        :return: Raw source row
        :rtype: tupel
        """

        return self.sheet[rowID]

    def getAbsoluteRowID(self, rowID):
        """
        Returns the absolute row ID from an relative row ID
                
        :param rowID: Relative row ID
        :type rowID: int
        :return: Absolute row ID
        :rtype: int
        """

        return rowID+self.headerRow+1

    def getRelativeRowID(self, rowID):
        """
        Returns the relative row ID from an absolute row ID
                
        :param rowID: Absolute row ID
        :type rowID: int
        :return: Absolute row ID
        :rtype: int
        """

        return rowID-self.headerRow-1

    def getColumnFromHeader(self, header):
        """
        Gets the column ID from an header
                
        :param header: Header ID
        :type header: string
        :return: Column ID or None if the header doesn't exist
        :rtype: int, None
        """

        if header.strip().lower() in self.header.keys():
            return self.header[header.strip().lower()][1].column
        else:
            return None

    def setRowValue(self, rowID, data):
        """
        Sets the data of a row
        
        Sets the data of the given row.
        Input is an dict of an key: value pair with headerName: newValue
        
        :param rowID: Relative row ID
        :type rowID: int
        :param data: New row data 
        :type data: dict
        """

        row = self.getRowFromID(rowID) #@TODO: implement checks
        for k,v in self.header.items():
            if v[0] in data.keys():
                row[v[1].col_idx-1].value = data[v[0]]
                row[v[1].col_idx-1].alignment = Alignment(wrapText=True)

    def getWarnings(self, clean=True):
        """
        Gets the warnings generated by getCellValue()
        
        Gets the warnings as list generated by getCellValue()
        
        :param clean: If set to True cleans the warnings list, defaults to True
        :param clean: bool, optional
        :return: List of all warnings
        :rtype: list
        """
        warnings = copy.copy(self.warnings)
        if clean is True:
            self.warnings = []
        return warnings

    def setDataValidatorForHeaderColumn(self, headerID, validatorType, formula, allowBlank=True):
        """
        Sets Datavalidators for specific header row
        
        Sets datavalidators for an header row with an special validation type and an formula
        (validator type is only list supported)
        
        :param headerID: ID of the header (name)
        :type headerID: string
        :param validatorType: validation type (only list supported at the moment)
        :type validatorType: string
        :param formula: formula of the validation
        :type formula: string
        :param allowBlank: If true blank cells are allowed, defaults to True
        :param allowBlank: bool, optional
        :raises KeyError: Raises if headerID not found in the headers
        :raises ValueError: Raises if there is already an validator for the headerID
        """
        headerID = str(headerID).strip().lower()
        if headerID not in self.header.keys():
            raise KeyError('Header ID %s not found' % headerID)
        if headerID in self.validators.keys():
            raise ValueError('Header ID %s has already an validator' % headerID)

        self.validators[headerID] = DataValidation(type=validatorType, formula1=formula, allow_blank=allowBlank)
        self.sheet.add_data_validation(self.validators[headerID])

    def addValidatorList(self, validatorList):
        """
        Adds an extra worksheet for data validation
        
        A worksheet with the name 'validatorLists' will be created to store lists for the validator
        
        @TODO: Check for existing validators = validatorList

        :param validatorList: list to write to validatorLists
        :type validatorList: list
        :return: coordinates of the generated list
        :rtype: string
        """

        if 'dropdown' not in self.doc:
            self.doc.create_sheet(title='dropdown')
        self.validatorListWorksheet = self.doc['dropdown']

        lastIndex = 1
        for x in self.validatorListWorksheet[1]:
            if x.value is not None:
                lastIndex += 1
        i = 1
        for v in validatorList:
            self.validatorListWorksheet.cell(row=i, column=lastIndex, value=v)
            i += 1
        return 'dropdown!$%s$%s:$%s$%s' % (cell.get_column_letter(lastIndex), 1, cell.get_column_letter(lastIndex), i-1)


    def addValidatorToCells(self, headerID):
        """
        Adds the registered validators to the workbook
        
        
        :param headerID: ID of the header
        :type headerID: string
        :raises KeyError: Raises if headerID not found in the headers
        :raises ValueError: Raises if there is no validator for the headerID
        """

        headerID = str(headerID).strip().lower()
        if headerID not in self.header.keys():
            raise KeyError('Header ID %s not found' % headerID)
        if headerID not in self.validators.keys():
            raise ValueError('Header ID %s has no validators' % headerID)
        for row in self.sheet:
            if row[0].row > self.headerRow:
                self.validators[headerID].add(row[self.header[headerID][1].col_idx-1])

    def addFilter(self):
        """
        Adds an filter in the header row        
        """
        self.sheet.auto_filter.ref = "A%s:%s%s" % (self.headerRow, cell.get_column_letter(self.lastColumn), self.headerRow)

    def setColumnWidth(self):
        """
        Sets the specified column width for the header column
        """
        for headerID in self.header:
            if self.header[headerID][6] != 0:
                self.sheet.column_dimensions[cell.get_column_letter(self.header[headerID][1].col_idx)].width = self.header[headerID][6]