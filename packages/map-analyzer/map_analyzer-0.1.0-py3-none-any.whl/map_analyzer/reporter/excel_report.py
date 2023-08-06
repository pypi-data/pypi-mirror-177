from openpyxl import Workbook
from openpyxl.styles import Alignment
from typing import List, Dict
from ..models import DetailItem, CoreGroup


class ExcelReporter:
    def __init__(self) -> None:
        self.wb = Workbook()

    def auto_width(self, worksheet):
        pass

    def write_title_row(self, sheet, title_row):
        for idx in range(0, len(title_row)):
            cell = sheet.cell(row=1, column=idx + 1)
            cell.value = title_row[idx]

    def write_cell(self, sheet, row, column, value, alignment = None):
        cell = sheet.cell(row = row, column=column) # type: cell
        cell.value = value
        if (alignment != None):
            cell.alignment = alignment

    def write_detail(self, items: List[DetailItem]):
        sheet = self.wb.create_sheet("Detail", 0)

        title_row = ["Name", "Group", "Core", "Section", "Type", "Size", "StartAddress", "EndAddress"]
        self.write_title_row(sheet, title_row)

        row = 2
        for item in items:
            self.write_cell(sheet, row, 1, item.name)
            self.write_cell(sheet, row, 2, item.group)
            self.write_cell(sheet, row, 3, item.core)
            self.write_cell(sheet, row, 4, item.section)
            self.write_cell(sheet, row, 5, item.type)
            self.write_cell(sheet, row, 6, item.size)
            self.write_cell(sheet, row, 7, "%x" % item.start_addr)
            self.write_cell(sheet, row, 8, "%x" % item.end_addr)
            row += 1

        self.auto_width(sheet)

    def write_summary(self, cores: Dict[str, CoreGroup]):
        sheet = self.wb.create_sheet("Summary", 1)

        title_row = ["Core", "Name", "Files", "text", "rodata", "bss", "data", "calibration"]
        self.write_title_row(sheet, title_row)

        row = 2
        for core_name in sorted(cores.keys()):
            for group_name in sorted(cores[core_name].groups.keys()):
                group = cores[core_name].groups[group_name]
                self.write_cell(sheet, row, 1, core_name)
                self.write_cell(sheet, row, 2, group_name)
                #self.write_cell(sheet, row, 2, group.files, Alignment(wrap_text=True, shrink_to_fit=True))
                self.write_cell(sheet, row, 4, group.text_total)
                self.write_cell(sheet, row, 5, group.rodata_total)
                self.write_cell(sheet, row, 6, group.bss_total)
                self.write_cell(sheet, row, 7, group.data_total)
                self.write_cell(sheet, row, 8, group.calib_total)
                row += 1

        self.auto_width(sheet)

    def save(self, name: str):
        self.wb.save(name)
